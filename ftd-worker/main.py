"""
FTD Parsing Worker — Minimal FastAPI for SiteDoc økonomi-modul.

Endepunkter:
  POST /parse/a-nota      — Parse A-nota Excel → strukturert JSON
  POST /parse/xml-budget   — Parse NS3459 XML → strukturert JSON
  POST /parse/mengdebeskrivelse — Parse mengdebeskrivelse PDF/Word → JSON
  GET  /health             — Helsesjekk
"""

from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import io

app = FastAPI(title="FTD Parser", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
async def health():
    return {"status": "ok", "version": "0.1.0"}


@app.post("/parse/a-nota")
async def parse_a_nota(file: UploadFile):
    """Parse A-nota Excel (xlsx/xls) → strukturert JSON med poster."""
    if not file.filename:
        raise HTTPException(400, "Mangler filnavn")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        raise HTTPException(400, f"Forventet Excel-fil, fikk .{ext}")

    content = await file.read()

    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(400, "Ingen aktive ark i filen")

        # Enkel parsing: les rader som poster
        headers = [cell.value for cell in ws[1]] if ws.max_row else []
        poster = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            post = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    post[str(headers[i])] = val
            poster.append(post)

        return {
            "status": "ok",
            "filename": file.filename,
            "antall_poster": len(poster),
            "headers": [h for h in headers if h],
            "poster": poster[:100],  # Begrens for respons
        }
    except ImportError:
        # openpyxl ikke installert — returner placeholder
        return {
            "status": "placeholder",
            "filename": file.filename,
            "message": "openpyxl ikke installert. Installer med: pip install openpyxl",
            "file_size": len(content),
        }


@app.post("/parse/xml-budget")
async def parse_xml_budget(file: UploadFile):
    """Parse NS3459 XML budsjett → strukturert JSON."""
    if not file.filename:
        raise HTTPException(400, "Mangler filnavn")

    content = await file.read()

    try:
        import xml.etree.ElementTree as ET

        root = ET.fromstring(content)

        # Finn alle poster (tilpass XPath til faktisk NS3459-skjema)
        poster = []
        for elem in root.iter():
            if elem.tag.endswith("Post") or elem.tag.endswith("Item"):
                post = {child.tag.split("}")[-1]: child.text for child in elem}
                poster.append(post)

        return {
            "status": "ok",
            "filename": file.filename,
            "root_tag": root.tag,
            "antall_poster": len(poster),
            "poster": poster[:100],
        }
    except Exception as e:
        raise HTTPException(400, f"Ugyldig XML: {e}")


@app.post("/parse/mengdebeskrivelse")
async def parse_mengdebeskrivelse(file: UploadFile):
    """Parse mengdebeskrivelse PDF/Word → tekstchunks med NS-koder."""
    if not file.filename:
        raise HTTPException(400, "Mangler filnavn")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    content = await file.read()

    if ext == "pdf":
        try:
            import pdfplumber

            chunks = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text() or ""
                    if text.strip():
                        chunks.append({
                            "page_number": i + 1,
                            "text": text,
                        })

            return {
                "status": "ok",
                "filename": file.filename,
                "page_count": len(chunks),
                "chunks": chunks,
            }
        except ImportError:
            return {
                "status": "placeholder",
                "filename": file.filename,
                "message": "pdfplumber ikke installert. Installer med: pip install pdfplumber",
                "file_size": len(content),
            }
    elif ext in ("docx", "doc"):
        return {
            "status": "placeholder",
            "filename": file.filename,
            "message": "Word-parsing kommer. Installer python-docx for .docx-støtte.",
            "file_size": len(content),
        }
    else:
        raise HTTPException(400, f"Ustøttet filtype: .{ext}")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8001)
